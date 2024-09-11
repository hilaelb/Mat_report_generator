"""
Export Matlab files to an Excel file.

One must be sure that there is an Mat_files folder filled with only the required Matlab files for the exporting at the
same folder as the python script.
"""
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from os import listdir
import os
import numpy as np
import pandas as pd
import scipy.io as sio
import time

__author__ = "Ophir Gruteke"
__copyright__ = "Copyright 2020, LAR"

__version__ = "1.0.0"
__maintainer__ = "Ophir Gruteke"
__email__ = "gruteke@post.bgu.ac.il"
__status__ = "Testing"

def timeit(method):
    """
    Time function by creating a decorator
    :param method: Function that we want to time.
    :return:
    """
    def timed(*args, **kw):
        ts = time.time()
        result = method(*args, **kw)
        te = time.time()
        if 'log_time' in kw:
            name = kw.get('log_name', method.__name__.upper())
            kw['log_time'][name] = int((te - ts) * 1000)
        else:
            print('%r  %2.2f ms' % \
                  (method.__name__, (te - ts) * 1000))
        return result
    return timed
@timeit
def fit_column_size(ws, data, headers):
    """
    Fit the column size by the longest word in the column.
    :param ws: Workbook sheet.
    :param data: Matrix of data.
    :param headers: List of headers name of each column.
    :return: None
    """
    column_widths = []
    for row in data:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]

    for i, cell in enumerate(headers):
        if len(column_widths) > i:
            if len(cell) > column_widths[i]:
                column_widths[i] = len(cell)
        else:
            column_widths += [len(cell)]

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width
@timeit
def add_data_to_sheet(ws,mat_data):
    """
    Add data from Matlab file to a Excel Sheet.
    :param ws: Workbook Sheet.
    :param mat_data: A file loaded by Scipy.
    :return: INT Maximum size of a column from Matlab data.
    """
    mat_contents = mat_data
    # Choose system id
    system_id = 2

    # Create Numpy arrays from Matlab file
    global_time = mat_contents['global_position_int'][0][system_id - 1][0]['time_boot_ms'][0][0] / 600000
    global_time = global_time - global_time[0]
    global_lat = mat_contents['global_position_int'][0][system_id - 1][0]['lat'][0][0] * 10**-7
    global_lon = mat_contents['global_position_int'][0][system_id - 1][0]['lon'][0][0] * 10**-7
    global_alt = mat_contents['global_position_int'][0][system_id - 1][0]['alt'][0][0]
    global_alt = np.around(global_alt, 2)
    global_depth = mat_contents['global_position_int'][0][system_id - 1][0]['depth'][0][0]
    global_depth = np.around(global_depth, 2)
    global_vx = mat_contents['global_position_int'][0][system_id - 1][0]['vx'][0][0]
    global_vx = np.around(global_vx, 3)
    global_vy = mat_contents['global_position_int'][0][system_id - 1][0]['vy'][0][0]
    global_vy = np.around(global_vy, 3)
    global_vz = mat_contents['global_position_int'][0][system_id - 1][0]['vz'][0][0]
    global_vz = np.around(global_vz, 3)
    global_hdg = mat_contents['global_position_int'][0][system_id - 1][0]['hdg'][0][0]
    global_hdg = np.around(global_hdg, 3)

    gps_time_usec = mat_contents['gps_raw_int'][0][0][0]['time_usec'][0][0] / 600000
    gps_time_usec = gps_time_usec-gps_time_usec[0]
    gps_time_fix_type = mat_contents['gps_raw_int'][0][0][0]['fix_type'][0][0]
    gps_time_lat = mat_contents['gps_raw_int'][0][0][0]['lat'][0][0] * 10**-7
    gps_time_lon = mat_contents['gps_raw_int'][0][0][0]['lon'][0][0] * 10**-7
    gps_time_vel = mat_contents['gps_raw_int'][0][0][0]['vel'][0][0] / 1000
    gps_time_vel = np.around(gps_time_vel, 3)

    svs_time = mat_contents['scaled_svs'][0][0][0]['time_boot_ms'][0][0] / 600000
    svs_time = svs_time - svs_time[0]
    svs_speed_of_sound = mat_contents['scaled_svs'][0][0][0]['Speed_of_Sound'][0][0]
    svs_speed_of_sound = np.around(svs_speed_of_sound, 1)
    svs_depth = mat_contents['scaled_svs'][0][0][0]['Depth'][0][0]
    svs_depth = np.around(svs_depth, 4)
    svs_temp = mat_contents['scaled_svs'][0][0][0]['Temperature'][0][0] / 100
    svs_temp = np.around(svs_temp, 2)

    pressure_time = mat_contents['scaled_pressure'][0][0][0]['time_boot_ms'][0][0] / 60000
    pressure_time = pressure_time - pressure_time[0]
    pressure_pressure = mat_contents['scaled_pressure'][0][0][0]['press_abs'][0][0]
    pressure_pressure = np.around(pressure_pressure, 4)
    pressure_temp = mat_contents['scaled_pressure'][0][0][0]['temperature'][0][0]
    pressure_temp = np.around(pressure_temp, 4)

    knot_velocity = (np.sqrt(np.square(global_vx)+np.square(global_vy))) * (1/0.5144)

    # Create space array to split between different tables.
    max_array_size = max(global_time.shape[0], gps_time_usec.shape[0], svs_time.shape[0], pressure_time.shape[0])
    space_array = np.array([''])
    space_array = np.pad(space_array, (0, max_array_size - 1), 'constant', constant_values=(''))

    array_list = [global_time, global_lat, global_lon, global_alt, global_depth, global_vx, global_vy, global_vz,
                  global_hdg, space_array, gps_time_usec, gps_time_fix_type, gps_time_lat, gps_time_lon, gps_time_vel,
                  space_array, svs_time, svs_speed_of_sound, svs_depth, svs_temp,
                  space_array, pressure_time, pressure_pressure, pressure_temp,space_array, global_time,knot_velocity]

    # Create even sized arrays by padding with zeros.
    for index, i in enumerate(array_list):
        # (0, max_array_size - i.shape[0]) add zeros only from the end.
        array_list[index] = np.pad(i, (0, max_array_size - i.shape[0]), 'constant', constant_values=(0))

    data = np.array(array_list).T
    df = pd.DataFrame(data)

    # Tables names, styles and alignment
    tables_name_list = ["Global Position",'','','','','','','','','',"GPS",'','','','','',
                        'SVS Data','','','','','Pressure Data','','','','Knot Velocity','']
    ws.append(tables_name_list)
    ws.append(['',''])
    ws['A1'].style = 'Title'
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['K1'].style = 'Title'
    ws['K1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['Q1'].style = 'Title'
    ws['Q1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['V1'].style = 'Title'
    ws['V1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['Z1'].style = 'Title'
    ws['Z1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.merge_cells('A1:I2')
    ws.merge_cells('K1:O2')
    ws.merge_cells('Q1:T2')
    ws.merge_cells('V1:X2')
    ws.merge_cells('Z1:AA2')

    # Add headers names
    headers = ['Time [min]', 'Latitude [degrees]', 'Longitude [degrees]',
               'Altitude [m]', 'Depth [m]', 'Vx [m/s]', 'Vy [m/s]', 'Vz [m/s]',
               'Hdg [deg]', '', 'Time [min]', 'Fix_type', 'Latitude [degrees]', 'Longitude [degrees]',
               'Vel [m/s]', '', 'Time [min]', 'Speed of Sound [m/s]', 'Depth [m]','Temp {C}',
               '', 'Time [min]', 'Press_abs [bar]', 'Temp [F]','','Time [min]','Knot Velocity [kn]']
    ws.append(headers)

    # Add rows from data frame to Excel sheet
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # Fit the column size after the sheet is filled.
    fit_column_size(ws, data, headers)
    return max_array_size

@timeit
def add_tablets_to_sheet(ws,max_array_size,sheet_number):
    """
    Add tables style to the sheet.
    :param ws: Workbook sheet.
    :param max_array_size: INT Length of maximun column.
    :param sheet_number: INT
    :return: None
    """
    tables_list = []
    # A dictionary with cell indexing for each table at the sheet.
    tablet_cell_dict = {0:"A3:I{d}".format(d=max_array_size+3),1:"K3:O{d}".format(d=max_array_size+3),
                        2:"Q3:T{d}".format(d=max_array_size+3),3:"V3:X{d}".format(d=max_array_size+3),4:"Z3:AA{d}".format(d=max_array_size+3)}
    for i in range((5*(sheet_number-1)),(5*sheet_number)):
        tab = Table(displayName="Table{d}".format(d=i), ref=tablet_cell_dict[i%5])
        # Add a default style with striped rows and banded columns
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tables_list.append(tab)

    # Add tablets to the sheet
    for i in tables_list:
        ws.add_table(i)
@timeit
def create_sheet_with_data(sheet, mat_data,sheet_number):
    """
    Create a complete sheet filled with tables.
    :param sheet: Workbook sheet.
    :param mat_data: A file loaded by Scipy.
    :param sheet_number: INT
    :return:
    """
    max_array_size = add_data_to_sheet(sheet, mat_data)
    add_tablets_to_sheet(sheet, max_array_size,sheet_number)

# Open a new Excel file
wb = Workbook()

# Get all Matlab files list

file_list = listdir(os.path.join(os.path.dirname(os.path.abspath(__file__)),'Mat_files'))
sheets_list = []
loaded_mat_files = []

# Create new sheets and load Matlab files.
for index, file in enumerate(file_list):
    sheets_list.append(wb.create_sheet("Experiment {d}".format(d=index+1)))   # Create a new sheet
    loaded_mat_files.append(sio.loadmat(os.path.join(os.path.join(os.path.dirname(os.path.abspath(__file__)),'Mat_files'),file)))   # Mat file loading

# Delete default sheet
del wb['Sheet']
sheet_number = 0
for index,sheet in enumerate(sheets_list):
    sheet_number += 1
    create_sheet_with_data(sheet , loaded_mat_files[index],sheet_number)

# Save and create the Excel file
print("Saving Workbook")
wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)),"Experiment.xlsx"))
print('Finished exporting')
